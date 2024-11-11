#!/usr/bin/python3

#   Copyright 2021-2022 Steve Strublic
#
#   This work is the personal property of Steve Strublic, and as such may not be
#   used, distributed, or modified without my express consent.

import os
import traceback
import datetime
import collections
import json
import re
import shutil

from elections import app, db
from elections import ADMINS

# Fetch event config.
import elections.events as events
from elections.events import EventConfig
from elections.ballotitems import ITEM_TYPES

from flask import redirect, render_template, url_for, request, session
from flask_login import current_user
from werkzeug.utils import secure_filename

import openpyxl, openpyxl.styles as styles, openpyxl.utils as utils

# Accepted True/False values.
VALID_TRUE_FALSE_VALUES = ['true', 'false', 'y', 'n', '1', '0', True, False]
VALID_TRUE_FALSE_VALUE_STR = "true/false, y/n or 1/0"
VALID_TRUE_VALUES = ['true', 'y', '1', True]

ALLOWED_EXTENSIONS = set(['xlsx'])
def allowed_file(filename):
	return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Export data versions.
# 1 - original
EXPORT_VERSION = 1

# All sheets to import/export.  This and SHEET_KEYS must be kept in sync to ensure data are processed correctly.
# This will have one entry per version.  Index 0 is empty.
ALL_SHEETS = [ [],
               ['events', 'ballotitems', 'candidates', 'voters', 'votes', 'vote_ballotid']
             ]

# Keys per sheet that we expect if the sheet is present.
SHEET_KEYS = {"events":               ['property', 'value'],
              "ballotitems":          ['itemid', 'type', 'name', 'description', 'positions', 'writeins'],
              "candidates":           ['id', 'itemid', 'firstname', 'lastname', 'fullname', 'writein'],
              "voters":               ['firstname', 'lastname', 'fullname', 'voteid', 'voted'],
              "votes":                ['itemid', 'ballotid', 'answer', 'commentary'],
              "vote_ballotid":        ['ballotid'],
             }

# Custom import parsing exceptions.
class ImportParseError(Exception):
    pass


# Retrieve the template file.
def downloadTemplate(user):
    try:
        event = current_user.event
        current_user.logger.info("Displaying: Fetch event template file")

        filename = 'event_data_template.xlsx'
        filepath = url_for('main_bp.exportfile', filename=filename)

        current_user.logger.info("Fetching event template file: Operation completed")

        return render_template('config/templatefile.html', user=user, admins=ADMINS[event.clubid],
                            filepath=filepath, filename=filename,
                            configdata=current_user.get_render_data())

    except Exception as e:
        current_user.logger.flashlog("Event Data Template failure", "Exception: %s" % str(e), propagate=True)
        current_user.logger.error("Unexpected exception:")
        current_user.logger.error(traceback.format_exc())

        # Redirect to the main page to display the exception and prevent recursive loops.
        return redirect(url_for('main_bp.index'))


# Export backup data to the user.
def exportData(user):
    try:
        event = current_user.event
        current_user.logger.info("Displaying: Export event data")

        if request.values.get('cancelbutton'):
            current_user.logger.flashlog(None, "Data export operation canceled.", 'info')
            return redirect(url_for('main_bp.exportdata'))

        saving = False
        if request.values.get('savebutton'):
            current_user.logger.info("Exporting data: Data export requested")
            saving = True

        filepath = None
        filename = None
        imagefilepath = None
        imagefilename = None

        if saving is True:
            fetchresults = request.values.get('results', False)
            if fetchresults is not False:
                fetchresults = True

            # Export the data to an excel file.  This will raise an exception upon failure.
            filepath, filename = buildExportFile(user, fetchresults)

        current_user.logger.info("Exporting data: Operation completed")

        return render_template('config/exportdata.html', user=user, admins=ADMINS[event.clubid],
                            filepath=filepath, filename=filename, imagefilepath=imagefilepath, imagefilename=imagefilename,
                            configdata=current_user.get_render_data())

    except Exception as e:
        current_user.logger.flashlog("Export Data failure", "Exception: %s" % str(e), propagate=True)
        current_user.logger.error("Unexpected exception:")
        current_user.logger.error(traceback.format_exc())

        # Redirect to the main page to display the exception and prevent recursive loops.
        return redirect(url_for('main_bp.index'))


# Build the sheet header row.
def build_header(ws, data):
    # The header is the first row.
    ws.append(data)
    row = ws["1:1"]
    for cell in row:
        cell.font = styles.Font(name='Calibri', bold=True)


# Fill the sheet with the data.
# The entry data must match the database field names exactly.
def fill_sheet(ws, data, entryfields):
    # Insert the header.
    build_header(ws, entryfields)

    # Store the maximum widths to pad the column width accordingly.
    maxwidths = [10] * len(entryfields)

    # Iterate through, adding the row's content and calculating the column width.
    for d in data:
        row = []
        for index, f in enumerate(entryfields):
            df = d.get(f, None)
            if df is None:
                row.append('')
            else:
                row.append(str(df))
                maxwidths[index] = max( maxwidths[index], len(str(df)) + 2 )

        ws.append(row)

    # Set the column width.
    for index, m in enumerate(maxwidths, start=1):
        ws.column_dimensions[utils.get_column_letter(index)].width = m


# Create the event information worksheet in the given workbook.
def createEventWorksheet(wb, table, event):
    ws = wb.active
    ws.title = table
    build_header(ws, ['property', 'value'])

    # Basic data.
    ws.append(['version', str(EXPORT_VERSION)])
    ws.append(['appversion', app.config.get('VERSION')])
    ws.append(['created', datetime.datetime.now()])
    ws.append(['title', event[0]['title']])
    ws.append(['eventdatetime', str(event[0]['eventdatetime'])])
    ws.append(['locked', str(event[0]['locked'])])

    ws["B4"].number_format = 'mm/dd/yyyy hh:mm:ss'
    ws.column_dimensions[utils.get_column_letter(1)].width = 20
    ws.column_dimensions[utils.get_column_letter(2)].width = len(str(ws["B4"].value))


# Build the export file.
def buildExportFile(user, fetchresults=False):
    try:
        event = current_user.event

        current_user.logger.debug("Exporting event data: Fetching event data", indent=1)

        # Read all class, entry, and vote data.
        outsql = []

        # Fetch app config data.  We can only save the app title.
        outsql.append('''SELECT *
                         FROM events
                         WHERE clubid='%d' AND eventid='%d';
                      ''' % (event.clubid, event.eventid))

        # Fetch ballot item data.
        outsql.append('''SELECT *
                         FROM ballotitems
                         WHERE clubid='%d' AND eventid='%d'
                         ORDER BY itemid ASC;
                      ''' % (event.clubid, event.eventid))

        # Fetch candidate data.
        outsql.append('''SELECT *
                         FROM candidates
                         WHERE clubid='%d' AND eventid='%d'
                         ORDER BY id ASC;
                      ''' % (event.clubid, event.eventid))

        # Fetch voter data.
        outsql.append('''SELECT *
                         FROM voters
                         WHERE clubid='%d' AND eventid='%d'
                         ORDER BY id ASC;
                      ''' % (event.clubid, event.eventid))

        # Fetch votes data.
        # This data does not need to be re-imported.
        outsql.append('''SELECT *
                         FROM votes
                         WHERE clubid='%d' AND eventid='%d'
                         ORDER BY id ASC;
                      ''' % (event.clubid, event.eventid))

        outsql.append('''SELECT *
                         FROM vote_ballotid
                         WHERE clubid='%d' AND eventid='%d';
                      ''' % (event.clubid, event.eventid))

        if fetchresults is True:
            current_user.logger.debug("Exporting event data: Including results", indent=1)

        # Fetch the data.
        _, data, _ = db.sql(outsql, handlekey=current_user.get_userid())

        current_user.logger.debug("Exporting event data: Building records", indent=1)

        # Data sets:
        # Put the data into a dict by table name.
        resultsdict = {}
        for index, table in enumerate(ALL_SHEETS[EXPORT_VERSION]):
            resultsdict[table] = data[index]

        # Build an excel file with tabs for:
        # - Non-importable data

        # Create a workbook.
        wb = openpyxl.Workbook()

        # Add the event worksheet.
        table = 'events'
        current_user.logger.debug("Exporting event data: Saving table '%s'" % table, indent=1)
        createEventWorksheet(wb, table, resultsdict[table])

        # Create all the other sheets.
        # Skip the event table, as we build that manually.
        # Use the 'newest' version of the sheets list as the currently implemented one.
        for table in ALL_SHEETS[EXPORT_VERSION]:
            if table != 'events':
                current_user.logger.debug("Exporting event data: Saving table '%s'" % table, indent=1)
                ws = wb.create_sheet(title=table)
                entryfields = get_sheet_keys(table, EXPORT_VERSION)
                fill_sheet(ws, resultsdict[table], entryfields)

        # Write the file to disk.  The filename is generic with the user's name postpended.
        filepath = os.path.join(os.getcwd(), app.config.get('EXPORT_DOWNLOAD_FOLDER'))
        if not os.path.exists(filepath):
            current_user.logger.info("Exporting event data: Created export data directory '%s'" % filepath, indent=1, propagate=True)
            os.makedirs(filepath)

        filename = 'election_%d_%d_export.xlsx' % (event.clubid, event.eventid)
        file = os.path.join(filepath, filename)

        current_user.logger.debug("Exporting event data: Saving data to file '%s'" % filename, indent=1)
        wb.save(filename=file)

        # Return the full file path and filename.
        current_user.logger.debug("Exporting event data: Saved data to file '%s'" % filename, indent=1)
        return url_for('main_bp.exportfile', filename=filename), filename

    except Exception as e:
        raise


# Reset to factory defaults.
def resetData(user):
    try:
        event = current_user.event

        current_user.logger.info("Displaying: Clear event data")

        # Clear any session flags.
        def clear_session_flags():
            for flag in ['reset', 'confirm']:
                if flag in session:
                    current_user.logger.debug("Clear event data: Clearing session flag '%s'" % flag, indent=1)
                    session.pop(flag, None)

        if request.values.get('cancelbutton'):
            current_user.logger.flashlog(None, "Clear event data operation canceled.", 'info')
            clear_session_flags()

            return redirect(url_for('main_bp.resetdata'))

        # The process is:
        # - Initiate via 'Submit'
        #   - Set a 'set' flag in the session
        # - Confirm via 'Confirm'
        #   - Set a 'confirm' flag in the session
        # - Execute the reset
        # - Confirm one more time as it's totally destructive
        reset_request = False
        confirm_request = False
        saving = False

        # Check if the event is locked.
        if current_user.event.locked is True:
            current_user.logger.flashlog("Clear Data failure", "This Event is locked and cannot clear Event data.")

        else:
            savebutton = request.values.get('savebutton', None)
            if savebutton is not None:

                if savebutton == 'reset':
                    current_user.logger.debug("Clear event data: Clear requested", propagate=True, indent=1)
                    session['reset'] = True

                    # Force reaquiring of confirmation.
                    if 'confirm' in session:
                        session.pop('confirm', None)

                    reset_request = True

                elif savebutton == 'confirm' and 'reset' in session:
                    current_user.logger.debug("Clear event data: Confirm requested", propagate=True, indent=1)
                    session['confirm'] = True

                    reset_request = True
                    confirm_request = True

                else:
                    if all(x in session for x in ['reset', 'confirm']):
                        saving = True
            else:
                # Force-clear session flags on fresh page load.
                clear_session_flags()

            if saving is True:
                current_user.logger.info("Clear event data: Initiating...", indent=1, propagate=True)

                # Force-clear session flags since we're done.
                clear_session_flags()
                reset_request = False
                confirm_request = False

                # Do the needful.
                err = reset_app_data(current_user.get_userid(), clear_config=False)
                if err is not None:
                    current_user.logger.flashlog("Clear Event Data failure", "Failed to reset event data:", highlight=True, propagate=True)
                    current_user.logger.flashlog("Clear Event Data failure", err, propagate=True)
                else:
                    current_user.logger.flashlog(None, "Clearing event data completed.", 'info', large=True, highlight=True, propagate=True)

                current_user.logger.info("Clear event data: Clear operation completed")

        return render_template('config/resetdata.html', user=user, admins=ADMINS[event.clubid],
                            reset_request=reset_request, confirm_request=confirm_request,
                            configdata=current_user.get_render_data())

    except Exception as e:
        current_user.logger.flashlog("Clear Event Data failure", "Exception: %s" % str(e), propagate=True)
        current_user.logger.error("Unexpected exception:")
        current_user.logger.error(traceback.format_exc())
        clear_session_flags()

        # Redirect to the main page to display the exception and prevent recursive loops.
        return redirect(url_for('main_bp.index'))


# Reset event voting data.
def restartEvent(user):
    try:
        event = current_user.event

        current_user.logger.info("Displaying: Restart event")

        # Clear any session flags.
        def clear_session_flags():
            for flag in ['restartevent', 'restartconfirm']:
                if flag in session:
                    current_user.logger.debug("Restart event: Clearing session flag '%s'" % flag, indent=1)
                    session.pop(flag, None)

        if request.values.get('cancelbutton'):
            current_user.logger.flashlog(None, "Restart event operation canceled.", 'info')
            clear_session_flags()

            return redirect(url_for('main_bp.restartevent'))

        # The process is:
        # - Initiate via 'Submit'
        #   - Set a 'set' flag in the session
        # - Confirm via 'Confirm'
        #   - Set a 'confirm' flag in the session
        # - Execute the reset
        # - Confirm one more time as it's totally destructive
        reset_request = False
        confirm_request = False
        saving = False

        # Check if the event is locked.
        if current_user.event.locked is True:
            current_user.logger.flashlog("Restart Event failure", "This Event is locked and cannot be restarted.")

        else:
            savebutton = request.values.get('savebutton', None)
            if savebutton is not None:

                if savebutton == 'reset':
                    current_user.logger.debug("Restart event: Clear requested", propagate=True, indent=1)
                    session['restartevent'] = True

                    # Force reaquiring of confirmation.
                    if 'eventconfirm' in session:
                        session.pop('restartconfirm', None)

                    reset_request = True

                elif savebutton == 'confirm' and 'restartevent' in session:
                    current_user.logger.debug("Restart event: Confirm requested", propagate=True, indent=1)
                    session['restartconfirm'] = True

                    reset_request = True
                    confirm_request = True

                else:
                    if all(x in session for x in ['restartevent', 'restartconfirm']):
                        saving = True
            else:
                # Force-clear session flags on fresh page load.
                clear_session_flags()

            if saving is True:
                current_user.logger.info("Restart event: Initiating...", indent=1, propagate=True)

                # Force-clear session flags since we're done.
                clear_session_flags()
                reset_request = False
                confirm_request = False

                # Do the needful.
                err = reset_app_data(current_user.get_userid(), clear_config=False, clear_only_votes=True)
                if err is not None:
                    current_user.logger.flashlog("Restart Event Data failure", "Failed to reset event voting data:", highlight=True, propagate=True)
                    current_user.logger.flashlog("Restart Event Data failure", err, propagate=True)
                else:
                    current_user.logger.flashlog(None, "Event restart completed.", 'info', large=True, highlight=True, propagate=True)

                current_user.logger.info("Restart event: Clear operation completed")

        return render_template('config/restartevent.html', user=user, admins=ADMINS[event.clubid],
                            reset_request=reset_request, confirm_request=confirm_request,
                            configdata=current_user.get_render_data())

    except Exception as e:
        current_user.logger.flashlog("Clear Event Data failure", "Exception: %s" % str(e), propagate=True)
        current_user.logger.error("Unexpected exception:")
        current_user.logger.error(traceback.format_exc())
        clear_session_flags()

        # Redirect to the main page to display the exception and prevent recursive loops.
        return redirect(url_for('main_bp.index'))


# Reset event data to defaults.
def reset_app_data(user, clear_config=True, clear_only_votes=False):
    try:
        event = current_user.event

        err = events.remove_event_data(user, event.clubid, event.eventid, clear_config=clear_config, votes_only=clear_only_votes)

        # On error to update the database, return and print out the error (like "System is in read only mode").
        if err is not None:
            return err
        else:
            # If specified, reset (reload) the event data.
            if clear_config is True:
                err = event.reset_config(user)
                if err is not None:
                    return err

            # If clearing everything, restore the default images.
            if clear_only_votes is False:
                # Copy the default images from storage.
                static_folder = os.path.join(os.getcwd(), app.config.get('STATIC_UPLOAD_FOLDER'))
                images_folder = os.path.join(os.getcwd(), app.config.get('IMAGES_UPLOAD_FOLDER'), str(event.clubid), str(event.eventid))

                if not os.path.exists(images_folder):
                    current_user.logger.debug("Creating images folder '%s" % images_folder, indent=1)
                    os.makedirs(images_folder)

                # If the config was cleared, put the original files back.  Clearing the config resets
                # the icon and homeimage file names to defaults.
                if clear_config is True:
                    for f in [event.icon, event.homeimage]:
                        try:
                            current_user.logger.debug("Restoring default image file '%s'" % f, indent=1)

                            srcfile = os.path.join(static_folder, f)
                            dstfile = os.path.join(images_folder, f)

                            if os.path.exists(dstfile):
                                os.remove(dstfile)

                            shutil.copy2(srcfile, dstfile)

                        except Exception as e:
                            current_user.logger.error("Restoring default image file: Failed to restore image file '%s" % f, indent=1, propagate=True)
                            raise Exception(str(e))

        return None

    except:
        raise


# Import event data and restore system content.
def importData(user):
    try:
        event = current_user.event

        current_user.logger.info("Displaying: Import event data")

        # Clear any session flags.
        def clear_session_flags():
            for flag in ['load', 'confirm', 'file', 'eventname']:
                if flag in session:
                    current_user.logger.debug("System data import: Clearing session flag '%s'" % flag, indent=1)
                    session.pop(flag, None)

        if request.values.get('cancelbutton'):
            current_user.logger.flashlog(None, "Data import operation canceled.", 'info')
            clear_session_flags()

            return redirect(url_for('main_bp.importdata'))

        # Read the file name from the form.
        filename = ''
        eventname = ''

        load_request = False
        confirm_request = False
        saving = False

        # Check if the event is locked.
        if current_user.event.locked is True:
            current_user.logger.flashlog("Import Data failure", "This Event is locked and cannot load Event data.")

        else:
            # If a file has previously been cached, fetch it.
            if 'file' in session:
                filepath = session['file']
                filename = filepath.split(os.sep)[-1]
            else:
                filepath = None
                filename = None


            # Generate default return action.
            def return_default():
                clear_session_flags()
                load_request = False
                confirm_request = False

                # If the file exists at the location, remove it before overwriting.
                if filepath is not None and os.path.exists(filepath):
                    current_user.logger.debug("System data import: Removing import file '%s'" % filename, indent=1)
                    os.remove(filepath)

                return render_template('config/importdata.html', user=user, admins=ADMINS[event.clubid],
                                    load_request=load_request, confirm_request=confirm_request, filename=None,
                                    configdata=current_user.get_render_data())

            # Find the event name from the event data.
            def getEventName(event):
                for e in event:
                    ep = e['property']
                    if ep == 'title':
                        return e['value']
                return ''


            savebutton = request.values.get('savebutton', None)
            if savebutton is not None:
                if savebutton == 'load':
                    if 'file' not in request.files:
                        current_user.logger.flashlog("Event Data Import failure", "No file information detected.")

                    # Get the filenam and cache in the session.
                    importfile = request.files['file']

                    if importfile.filename == '':
                        current_user.logger.flashlog("Event Data Import failure", "No file was provided.")
                        return return_default()

                    if not allowed_file(importfile.filename):
                        current_user.logger.flashlog("Event Data Import failure", "Unsupported file type (Valid types: %s)." % ', '.join(ALLOWED_EXTENSIONS))
                        return return_default()

                    current_user.logger.info("System data import: Load requested for file '%s'" % importfile.filename, propagate=True, indent=1)

                    # Cache the file to disk and record the path.
                    basepath = os.path.join(os.getcwd(), app.config.get('IMPORT_UPLOAD_FOLDER'))
                    if not os.path.exists(basepath):
                        current_user.logger.debug("System data import: Creating import file directory '%s'" % basepath, indent=1)
                        os.makedirs(basepath)

                    # If the file exists at the location, remove it before overwriting.
                    filename = secure_filename(importfile.filename)
                    filepath = os.path.join(basepath, filename)

                    if os.path.exists(filepath):
                        current_user.logger.info("System data import: Overriding import file '%s'" % filename, indent=1, propagate=True)
                        os.remove(filepath)

                    current_user.logger.debug("System data import: Saving uploaded file '%s'" % filename, indent=1)
                    importfile.save(filepath)

                    session['load'] = True
                    session['file'] = filepath

                    # Force reaquiring of confirmation.
                    if 'confirm' in session:
                        session.pop('confirm', None)

                    load_request = True
                    eventname = '<Validation Required>'

                elif savebutton == 'validate' and 'load' in session:
                    current_user.logger.debug("System data import: Validate requested", propagate=True, indent=1)

                    # Read the data, and then validate it if it read successfully.
                    data, _, url = validateImportData(filepath, validating=True)
                    if url is not None:
                        return url

                    load_request = True
                    eventname = getEventName(data['events'])

                    # Stash the name as validated.
                    session['eventname'] = eventname

                elif savebutton == 'confirm' and 'load' in session:
                    current_user.logger.debug("System data import: Confirm requested", propagate=True, indent=1)
                    session['confirm'] = True

                    load_request = True
                    confirm_request = True

                    eventname = session.get('eventname', '')

                else:
                    if all(x in session for x in ['load', 'confirm']):
                        saving = True
            else:
                # Force-clear session flags on fresh page load.
                clear_session_flags()

            if saving is True:
                current_user.logger.info("Importing event data: Loading from file '%s'" % filename, indent=1, propagate=True)

                # Force-clear remaining session flags since we're done.
                clear_session_flags()
                load_request = False
                confirm_request = False

                # Read and validate the file content.
                data, validationdata, url = validateImportData(filepath)
                if url is not None:
                    return url

                eventname = getEventName(data['events'])

                # Import all available data.
                current_user.logger.info("Importing event data: Loading data...", indent=1, propagate=True)

                url = importValidatedData(data, validationdata, user)
                if url is not None:
                    return url

                current_user.logger.flashlog(None, "Import Successful.", level='info', highlight=True, large=True, propagate=True)

                current_user.logger.info("Import event data: Import operation completed")
                return redirect(url_for('main_bp.index'))

        return render_template('config/importdata.html', user=user, admins=ADMINS[event.clubid],
                            load_request=load_request, confirm_request=confirm_request, filename=filename, eventname=eventname,
                            configdata=current_user.get_render_data())

    except Exception as e:
        current_user.logger.flashlog("Event Data Import failure", "Exception: %s" % str(e), propagate=True)
        current_user.logger.error("Unexpected exception:")
        current_user.logger.error(traceback.format_exc())

        # Redirect to the main page to display the exception and prevent recursive loops.
        return redirect(url_for('main_bp.index'))


# Read the data into a dict.
def readImportData(filepath):
    current_user.logger.info("Import event data: Reading data", propagate=True)

    # Create an 'ordered' dict.
    data = collections.OrderedDict()

    # The file must exist.
    if filepath is None or not os.path.exists(filepath):
        raise IOError("File not found")

    try:
        # The file must be an excel file.
        wb = openpyxl.load_workbook(filename=filepath, data_only=True)

    except Exception as e:
        # The errors are not clear here, so we catch it and raise an IOError for ourselves.
        raise IOError("Invalid file format")

    # The bare minimum required is a page for config (version), classes and entries.
    # But we'll take what we can get.
    errors = []
    for sheet in ['events']:
        if sheet not in wb.sheetnames:
            errors.append("Missing required sheet '%s'" % sheet)

    if len(errors) > 0:
        for error in errors:
            current_user.logger.error("Reading import data: %s" % error)
        raise ImportParseError(errors)


    # Read a worksheet into a list of dict entries.
    def read_sheet(ws, sheetname, sheetversion):
        sheetdata = []
        header_row = None

        sheet_keys = get_sheet_keys(sheetname, sheetversion)

        # Row iterator to emulate a list per-row.
        def iter_rows(ws):
            for row in ws.iter_rows():
                yield [cell.value for cell in row]

        rows = list(iter_rows(ws))

        # Get the header row.
        def get_header_row():
            # We expect the header row to be row 1.
            # A quick check is to look at the first key and see if
            # it is in the row.
            row = rows[0]
            if sheet_keys[0] in row:
                header_row = row
            else:
                header_row = None

            return header_row

        # Look for all sheet keys.
        def check_sheet_keys():
            key_errors = []
            # We want to know that all the expected keys in the sheet
            # are in the header.  Any extra keys will be ignored.
            if not all(k in header_row for k in sheet_keys):
                missing_keys = list(set(sheet_keys).difference(header_row))
                key_errors.append("Sheet '%s': Missing columns '%s'" % (sheetname, ','.join(missing_keys)))

            # Also check for duplicate keys.
            counts = collections.Counter(header_row)
            for c in counts:
                if counts[c] != 1:
                    key_errors.append("Sheet '%s': Duplicate column '%s'" % (sheetname, c))

            # Return any errors, or None for A-OK.
            if len(key_errors) > 0:
                return key_errors

            # We are guaranteed that all keys are present and not duplicated.
            return None

        # Check that there is a header row.
        header_row = get_header_row()
        if header_row is None:
            return None, ["Sheet '%s': Missing header row" % sheet]

        # Check that the keys we need are all present and unique.
        key_errors = check_sheet_keys()
        if key_errors is not None:
            return None, key_errors

        # Get the row data.
        for row in rows[1:]:
            # Read the row and assign the content to the dict.
            # The keys may be out-of-order, so we need the index
            # of each from the header row.  We force to string
            # and strip them for tidiness.
            rowdata = {}
            for key in sheet_keys:
                colindex = header_row.index(key)

                if row[colindex] is None:
                    rowdata[key] = None
                else:
                    rowdata[key] = str(row[colindex]).strip()

            # If the row is blank, skip it.
            add = False
            for key in sheet_keys:
                if rowdata[key] is not None and len(rowdata[key]) != 0:
                    add = True
                    break

            if add is True:
                sheetdata.append(rowdata)

        return sheetdata, None


    # For file load, all we want are the sheets that are available and to read them
    # into a dict.  Extra sheets will be flagged but ignored.

    # We need the sheets for the version that was specified.
    # This comes from the 'events' sheet, which must always exist and have a version.
    sheet = 'events'
    sheetversion = 0

    ws = wb[sheet]

    # We always read the newest sheet for the events table.  The table is backward compaible.
    sheetdata, sheeterrors = read_sheet(ws, sheet, EXPORT_VERSION)
    if sheeterrors is not None:
        errors.extend(sheeterrors)
    else:
        # Fetch the version from the events sheet and validate.
        try:
            for s in sheetdata:
                p = s.get('property')
                if p == 'version':
                    sheetversion = int(s.get('value'))
                    break

        except:
            errors.append(["Events: Version field must be an integer"])

        # Did we find it?
        if sheetversion == 0:
            errors.append(["Events: Missing or incorrect 'version' field"])
        else:
            # Check version.
            if sheetversion < 1 or sheetversion > EXPORT_VERSION:
                errors.append(["Events: Invalid version '%d' detected" % sheetversion])
            else:
                data[sheet] = sheetdata

    # Can't read anything else if the version is not valid.
    if len(errors) == 0:
        # Read all the other sheets.
        for sheet in ALL_SHEETS[sheetversion]:
            if sheet != 'events' and sheet in wb.sheetnames:
                ws = wb[sheet]
                sheetdata, sheeterrors = read_sheet(ws, sheet, sheetversion)
                if sheeterrors is not None:
                    errors.extend(sheeterrors)
                else:
                    data[sheet] = sheetdata

    # If any errors were found during the read process, raise them now.
    if len(errors) > 0:
        for error in errors:
            current_user.logger.error("Reading import data: %s" % error)

        raise ImportParseError(errors)

    return data, sheetversion


# Validate the import data.
# If called to validate-only, it affects the output messages on failure.
# Once called to validate, we're assured the data will be good - but we'll validate again
# when we do the actual import.
# We'll return the data and any data we used to validate it, as it might be useful when importing the data.
# On error, we'll return the redirect URL.
def validateImportData(filepath, validating=False):
    msg = "Import"
    if validating is True:
        msg = "Validation"

    # Read the import data into a dict.
    try:
        data, sheetversion = readImportData(filepath)

        current_user.logger.info("Importing event data: Converting data...", indent=1, propagate=True)
        data = convertImportedData(data, sheetversion)

    except IOError as i:
        # An IOError means the file wasn't a valid Excel file.
        current_user.logger.flashlog("Event Data Import failure", "Data %s failed:" % msg, highlight=True, large=True, propagate=True)
        current_user.logger.flashlog("Event Data Import failure", "%s" % str(i), indent=True, propagate=True)

        return None, None, redirect(url_for('main_bp.importdata'))

    except ImportParseError as ip:
        # An ImportParseError means we didn't like something in the file.
        # This will always come back as a list of messages so we can log them to the user.
        current_user.logger.flashlog("Event Data Import failure", "Data %s failed: Invalid file content:" % msg, highlight=True, large=True, propagate=True)
        errs = json.loads(str(ip))
        for err in errs:
            current_user.logger.flashlog("Event Data Import failure", "%s" % err, indent=True, propagate=True)

        return None, None, redirect(url_for('main_bp.importdata'))

    except Exception as e:
        # Something else.
        current_user.logger.flashlog("Event Data Import failure", "Data %s failed:" % msg, highlight=True, large=True, propagate=True)
        current_user.logger.flashlog("Event Data Import failure", "%s" % str(e), propagate=True, indent=True)

        return None, None, redirect(url_for('main_bp.importdata'))

    event = current_user.event

    # If the imported data has certain worksheets with content, then the data is a backup/restore and not a bulk import of new entries and classes.
    restoring_backup = False
    if any(sheet in list(data) for sheet in ['vote_ballotid', 'votes']):
        current_user.logger.info("Restoring a backup for event %d" % event.eventid, indent=1, propagate=True)
        restoring_backup = True

    # Now we have the data with all expected fields and the minimum required sheets.
    # Let's verify the content.
    warnings = []

    # Seed the return data with the club and event ID we expect as well as the sheet input version.
    returndata = {'clubid': event.clubid,
                  'eventid': event.eventid,
                  'version': sheetversion,
                  'restoring_backup': restoring_backup,
                  'max_ballotid': 0
                 }

    # Build a table of validator methods.  Each takes in the data set and data fetched by previous validators.
    # It returns the same parameters: errors, warnings, returndata (data read from and needed by other things).
    # Order is important as other validators may need things fetched by prior validators.
    validation_table = [
                        validateEventConfig,
                        validateBallotItems,
                        validateCandidates,
                        validateVoters,
                        validateVotes
                       ]

    # Iterate through the table, calling each validator.
    # On error, dump the errors to the session and return.
    # Warnings and returndata are aggregated and handled at the end.
    current_user.logger.info("Import event data: Validating data", indent=1, propagate=True)

    for v in validation_table:
        v_errors, v_warnings, v_returndata = v(data, returndata)

        # For any errors, return now.
        if len(v_errors) > 0:
            current_user.logger.flashlog("Event Data Import failure", "Data %s failed:" % msg, highlight=True, large=True, propagate=True)

            # Limit to 25 to prevent overloading the session.  That's a lot of errors.
            for e in v_errors[0:24]:
                current_user.logger.flashlog("Event Data Import Failure", e, indent=True, propagate=True)

            if len(v_errors) > 25:
                current_user.logger.flashlog("Event Data Import Failure", "Too many errors to continue", indent=True, propagate=True)

            return None, None, redirect(url_for('main_bp.importdata'))

        # Append any warnings.
        if len(v_warnings) > 0:
            warnings.extend(v_warnings)

        # Grab any return data.
        if len(v_returndata) > 0:
            returndata.update(v_returndata)

    # ==============================================
    # We made it!
    # ==============================================
    current_user.logger.flashlog(None, "Validation Successful.", highlight=True, large=True, level='info', propagate=True)

    # Dump any warnings.
    if len(warnings) > 0:
        current_user.logger.flashlog(None, "The following warnings were encountered:", propagate=True)

        for w in warnings:
            current_user.logger.flashlog(None, w, indent=True, propagate=True)

    # Return data set, validation data set, no error URL.
    return data, returndata, None


# Little helper to consolidate checking for True/False.
def valid_true_false_value(value):
    # Value must be 'True/False', 'Y/N', or '1/0'.
    try:
        return value.lower() in VALID_TRUE_FALSE_VALUES
    except:
        return False


# Little helper for checking if a value is True.
def value_is_true(value):
    try:
        if valid_true_false_value(value):
            return value.lower() in VALID_TRUE_VALUES
    except:
        return False


# Little helper for checking if a value is in a range.
def value_in_range(value, low, high):
    try:
        # Must be an integer, and within the range low-high.
        # Add 1 to the high value as that's how Python works.
        value = int(value)
        return value in range(low, (high + 1))
    except:
        return False

# Little helper to get True/False for a given input value.
def true_false_value(value):
    try:
        if value.lower() in VALID_TRUE_VALUES:
            return True
    except:
        pass

    return False

# Little helper to get the sheet keys for the given import data version.
def get_sheet_keys(table, sheetversion):
    keys = SHEET_KEYS[table]

    # If the first element in the keys list is a list, then this is a list of lists.
    # We want the one that corresponds to the version we imported.
    try:
        if type(keys[0]) is list:
            keys = keys[sheetversion]

    except:
        raise

    return keys


# Validate event data.
def validateEventConfig(data, appdata):
    current_user.logger.debug("Validating event data", indent=1)

    # ==============================================
    # Read the event config data to fetch the version and check the title.
    # ==============================================
    errors = []
    warnings = []
    returndata = {}

    table = 'events'

    dataversion = 0
    eventdatetime = ''

    # This table is required.
    if table not in data:
        errors.append("Missing required table '%s'" % table)
        return errors, warnings, returndata

    # The events table is unique in that it is a property-value pairs list.
    event = data[table]

    # Look for the 'title' and 'version' properties.
    for index, a in enumerate(event):
        p = a['property']

        if p in ['title', 'version', 'eventdatetime'] and a['value'] is None:
            errors.append("EventConfig: Row %d: Missing value for property '%s'" % (index, p))

        else:
            # Check values.
            if p in  ['version']:
                try:
                    value = int(a['value'])

                    # Version has been prevalidated when the import data was read.
                    if p == 'version':
                        dataversion = value

                except:
                    errors.append("EventConfig: Row %d: Value for column '%s' must be an integer" % (index, p))

            elif p == 'eventdatetime':
                # Must be a datetime-local (ISO8601) formatted value.
                # The seconds value is optional and we'll strip it when we save it.
                if not re.match(r'^\d{4}-(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])T? ?(0[0-9]|1[0-9]|2[0-3]):([0-5][0-9]):?([0-5][0-9])?$', a['value']):
                    errors.append("EventConfig: Row %d: Value for column '%s' must be a valid date in format YYYY-MM-DDTHH:MM" % (index, p))
                else:
                    # Drop the last part (seconds) if included.
                    # If the 'T' is missing (LibreOffice does this for date format), restore it.
                    if len(a['value'].split(':')) == 3:
                        eventdatetime = ':'.join(a['value'].split(':')[:2]).replace(' ', 'T')
                    else:
                        eventdatetime = a['value'].replace(' ', 'T')

    # Missing version is bad.  We need this to vet if what comes in is supported.
    if dataversion == 0:
        errors.append("EventConfig: Missing property 'version'")
    else:
        # Check that the version is supported.
        supported_versions = app.config.get('IMPORT_SUPPORTED_VERSIONS')
        if dataversion not in supported_versions:
            errors.append("EventConfig: Unsupported version '%d' (supported versions: %s)" % (dataversion, ','.join(str(x) for x in supported_versions)))

    # Put the data we have in the return data dict.
    returndata['dataversion'] = dataversion
    returndata['eventdatetime'] = eventdatetime

    return errors, warnings, returndata


# Validate the ballot ID information for class choice ballots.
def validateBallotID(data, appdata):
    current_user.logger.debug("Validating ballot ID", indent=1)

    errors = []
    warnings = []
    returndata = {}

    table = 'vote_ballotid'

    # Ballot ID data is optional
    if table in data:
        vote_ballotid = data[table]

        # There must be only one entry in this table.  But it can be empty.
        rowcount = len(vote_ballotid)
        if rowcount > 1:
            errors.append("Vote Ballot IDs: Row count must be 1 (rows: %d)" % rowcount)
        else:
            # We start at '2' because row 1 is the header row, previously verified.
            # The data is in order in the ordered dict, so it represents row value.
            for index, i in enumerate(vote_ballotid, start=2):
                # Verify all fields have appropriate values.
                keys = get_sheet_keys(table, appdata['version'])
                for key in keys:
                    if i.get(key, None) is None:
                        errors.append("Vote Ballot IDs: Row %d: Missing value for column '%s'" % (index, key))
                        continue

                    value = i.get(key)

                    try:
                        # Value must be an integer in a range.
                        value = int(value)
                        if value < 0:
                            errors.append("Vote Ballot IDs: Row %d: Value for column '%s' is invalid (%d)" % (index, key, value))

                    except:
                        errors.append("Vote Ballot IDs: Row %d: Value for column '%s' must be an integer" % (index, key))

            # Check that the one and only entry is equal to or greater than the last used ballot ID.
            if rowcount == 1 and len(errors) == 0:

                # This will have been accumulated / maxed by all validators of things that use the vote ballotid counter.
                max_ballotid = appdata.get('max_ballotid', 0)

                # We have exactly one record, so fetch its vote_ballotid.
                row = 0
                vote_ballotid = int(vote_ballotid[row]['ballotid'])

                # The recorded value must be equal to or above the highest used ballot value.
                # They might not match if the most recent ballot(s) were removed from the event.
                if vote_ballotid < max_ballotid:
                    errors.append("Vote Ballot IDs: Next ballot ID %d does not align with highest ballot ID %d" % (vote_ballotid, max_ballotid))

    else:
        current_user.logger.debug("No data provided: skipping", indent=2)

    return errors, warnings, returndata


def validateBallotItems(data, appdata):
    current_user.logger.debug("Validating ballot item data", indent=1)

    errors = []
    warnings = []
    returndata = {}

    table = 'ballotitems'

    # This table is required.
    if table not in data:
        errors.append("Missing required table '%s'" % table)
        return errors, warnings, returndata

    ballotitems = data[table]

    # We start at '2' because row 1 is the header row, previously verified.
    # The data is in order in the ordered dict, so it represents row value.
    for index, e in enumerate(ballotitems, start=2):
        # Verify all fields have appropriate values.
        keys = get_sheet_keys(table, appdata['version'])
        for key in keys:
            # Any required key must have a value.
            value = e.get(key, None)

            if value is None:
                errors.append("Ballot Items: Row %d: Missing value for column '%s'" % (index, key))
                continue

            # Check type and range where appropriate.
            if value is not None:
                if key in ['itemid', 'type', 'positions']:
                    try:
                        value = int(value)

                        # Class end entry IDs must be positive values.
                        if key in ['itemid', 'positions']:
                            if value < 0:
                                errors.append("Ballot Items: Row %d: Value for column '%s' is invalid (%d)" % (index, key, value))

                        elif key in ['type']:
                            typecount = 2
                            if not value_in_range(value, 1, 2):
                                errors.append("Ballot Items: Row %d: Value for column '%s' is invalid (%d) (range: 1 - %d)" % (index, key, value, typecount))

                    except:
                        errors.append("Ballot Items: Row %d: Value for column '%s' must be an integer" % (index, key))

                elif value in ['writeins']:
                    if not valid_true_false_value(value):
                        errors.append("Ballot Items: Row %d: Value for column '%s' must be %s" % (index, key, VALID_TRUE_FALSE_VALUES))

    # Validate the data itself.
    if len(errors) == 0:
        itemlist = []
        namelist = []
        desclist = []

        for b in ballotitems:
            itemlist.append(b['itemid'])
            namelist.append(b['name'])
            desclist.append(b['description'])

        # Check that item IDs, names and descriptions are not duplicated.
        for index, itemid in enumerate(itemlist, start=2):
            if itemlist.count(itemid) > 1:
                errors.append("Ballot Items: Row %d: Item ID %d is duplicated" % (index, itemid))

        for index, name in enumerate(namelist, start=2):
            if namelist.count(name) > 1:
                errors.append("Ballot Items: Row %d: Name '%s' is duplicated" % (index, name))

        for index, name in enumerate(desclist, start=2):
            if desclist.count(name) > 1:
                errors.append("Ballot Items: Row %d: Description is duplicated" % index)

    # If all checks out, save the ballot items as a dict for future validation.
    if len(errors) == 0:
        items = {}
        for b in ballotitems:
            item = {'itemid': b['itemid'],
                    'type': b['type'],
                    'name': b['name'],
                    'description': b['description'],
                    'positions': b['positions'],
                    'writeins': b['writeins']
                   }
            items[b['itemid']] = item

        returndata['ballotitems'] = items

    return errors, warnings, returndata


def validateCandidates(data, appdata):
    current_user.logger.debug("Validating candidate data", indent=1)

    errors = []
    warnings = []
    returndata = {}

    table = 'candidates'

    # This table is required.
    if table not in data:
        errors.append("Missing required table '%s'" % table)
        return errors, warnings, returndata

    candidates = data[table]

    # We start at '2' because row 1 is the header row, previously verified.
    # The data is in order in the ordered dict, so it represents row value.
    for index, e in enumerate(candidates, start=2):
        # Verify all fields have appropriate values.
        keys = get_sheet_keys(table, appdata['version'])
        for key in keys:
            # Any required key must have a value.
            value = e.get(key, None)

            if value is None:
                errors.append("Candidates: Row %d: Missing value for column '%s'" % (index, key))
                continue

            # Check type and range where appropriate.
            if value is not None:
                if key in ['itemid']:
                    try:
                        value = int(value)

                        # Class end entry IDs must be positive values.
                        if key in ['itemid']:
                            if value < 0:
                                errors.append("Candidates: Row %d: Value for column '%s' is invalid (%d)" % (index, key, value))

                    except:
                        errors.append("Candidates: Row %d: Value for column '%s' must be an integer" % (index, key))

                elif value in ['writein']:
                    if not valid_true_false_value(value):
                        errors.append("Candidates: Row %d: Value for column '%s' must be %s" % (index, key, VALID_TRUE_FALSE_VALUES))

    # Validate the data itself.
    if len(errors) == 0:
        idlist = []
        namelist = []

        ballotitems = appdata['ballotitems']
        for c in candidates:
            if c['itemid'] not in ballotitems.keys():
                errors.append("Candidates: Row %d: Ballot item id '%d' was not found" % (index, c['itemid']))

        # Verify the first and last names match the full name.
        for c in candidates:
            fullname = c['fullname']

            first, last = fullname.split(' ', 1)
            if first != c['firstname']:
                errors.append("Candidates: Row %d: First name '%s' does not match full name" % (index, c['firstname']))

            if last != c['lastname']:
                errors.append("Candidates: Row %d: Last name '%s' does not match full name" % (index, c['lastname']))

        for c in candidates:
            idlist.append(c['id'])
            namelist.append(c['fullname'])

        for index, name in enumerate(namelist, start=2):
            if idlist.count(name) > 1:
                errors.append("Candidates: Row %d: ID '%d' is duplicated" % (index, id))

        for index, name in enumerate(namelist, start=2):
            if namelist.count(name) > 1:
                errors.append("Candidates: Row %d: Name '%s' is duplicated" % (index, name))

    # If all checks out, save the ballot items as a dict for future validation.
    if len(errors) == 0:
        for c in candidates:
            itemid = c['itemid']
            candidateid = c['id']

            item = {'firstname': c['firstname'],
                    'lastname': c['lastname'],
                    'fullname': c['fullname'],
                    'writein': c['writein']
                   }

            if 'candidates' not in ballotitems[itemid]:
                ballotitems[itemid]['candidates'] = {}

            ballotitems[itemid]['candidates'][candidateid] = item

    return errors, warnings, returndata


def validateVoters(data, appdata):
    current_user.logger.debug("Validating voter data", indent=1)

    errors = []
    warnings = []
    returndata = {}

    table = 'voters'

    # This table is required.
    if table in data:
        voters = data[table]

        # We start at '2' because row 1 is the header row, previously verified.
        # The data is in order in the ordered dict, so it represents row value.
        for index, e in enumerate(voters, start=2):
            # Verify all fields have appropriate values.
            keys = get_sheet_keys(table, appdata['version'])
            for key in keys:
                # Any required key must have a value.
                value = e.get(key, None)

                if value is None:
                    errors.append("Voters: Row %d: Missing value for column '%s'" % (index, key))
                    continue

                # Check type and range where appropriate.
                if value is not None:
                    if value in ['voted']:
                        if not valid_true_false_value(value):
                            errors.append("Voters: Row %d: Value for column '%s' must be %s" % (index, key, VALID_TRUE_FALSE_VALUES))

        # Validate the data itself.
        if len(errors) == 0:
            namelist = []

            # Verify the first and last names match the full name.
            for v in voters:
                fullname = v['fullname']
                first, last = fullname.split(' ', 1)
                if first != v['firstname']:
                    errors.append("Voters: Row %d: First name '%s' does not match full name" % (index, v['firstname']))

                if last != v['lastname']:
                    errors.append("Voters: Row %d: Last name '%s' does not match full name" % (index, v['lastname']))

            for v in voters:
                namelist.append(v['fullname'])

            for index, name in enumerate(namelist, start=2):
                if namelist.count(name) > 1:
                    errors.append("Voters: Row %d: Name '%s' is duplicated" % (index, name))

    return errors, warnings, returndata


def validateVotes(data, appdata):
    errors = []
    warnings = []
    returndata = {}

    table = 'votes'

    # This table is not required.
    if table in data:
        votes = data[table]
        current_user.logger.debug("Validating vote data", indent=1)

        # We start at '2' because row 1 is the header row, previously verified.
        # The data is in order in the ordered dict, so it represents row value.
        for index, e in enumerate(votes, start=2):
            # Verify all fields have appropriate values.
            keys = get_sheet_keys(table, appdata['version'])
            for key in keys:
                required = True
                if key in ['commentary']:
                    required = False

                # Any required key must have a value.
                value = e.get(key, None)

                if required is True and value is None:
                    errors.append("Votes: Row %d: Missing value for column '%s'" % (index, key))
                    continue

                # Check type and range where appropriate.
                if value is not None:
                    if key in ['ballotid', 'itemid', 'answer']:
                        try:
                            value = int(value)

                            # Class end entry IDs must be positive values.
                            if key in ['ballotid', 'itemid', 'answer']:
                                if value < 0:
                                    errors.append("Votes: Row %d: Value for column '%s' is invalid (%d)" % (index, key, value))

                        except:
                            errors.append("Votes: Row %d: Value for column '%s' must be an integer" % (index, key))

        # Validate the data itself.
        if len(errors) == 0:
            ballotitems = appdata['ballotitems']

            for b in ballotitems:
                ballotitem = ballotitems[b]
                itemid = ballotitem['itemid']

            counts = {}
            for v in votes:
                itemid = v['itemid']
                ballotid = v['ballotid']

                if itemid not in list(ballotitems.keys()):
                    errors.append("Votes: Row %d: Ballot item ID %d is invalid" % (index, itemid))
                else:
                    answer = v['answer']
                    itemtype = int(ballotitems[itemid]['type'])

                    if itemtype == ITEM_TYPES.CONTEST.value:
                        candidates = ballotitems[itemid]['candidates']
                        if answer not in list(candidates.keys()):
                            errors.append("Votes: Row %d: Answer '%s' does not match a candidate for contest ballot item %s" % (index, answer, itemid))

                    elif itemtype == ITEM_TYPES.QUESTION.value:
                        if answer is not None and int(answer) not in range(0, 2):
                            errors.append("Votes: Row %d: Answer '%s' is invalid for question ballot item %s" % (index, answer, itemid))

                if itemid not in counts:
                    counts[itemid] = {}

                if ballotid not in counts[itemid]:
                    counts[itemid][ballotid] = 0

                counts[itemid][ballotid] += 1

            for b in ballotitems:
                ballotitem = ballotitems[b]
                itemid = ballotitem['itemid']
                itemtype = int(ballotitem['type'])

                if itemtype == ITEM_TYPES.CONTEST.value:
                    positions = int(ballotitem['positions'])

                    for ballotid in counts[itemid]:
                        ballotcount = counts[itemid][ballotid]
                        if ballotcount > positions:
                            errors.append("Votes: Ballot Item %s has more votes (%d) than positions (%d)" % itemid, (c, positions))

                elif itemtype == ITEM_TYPES.QUESTION.value:
                    for ballotid in counts[itemid]:
                        ballotcount = counts[itemid][ballotid]
                        if ballotcount > 1:
                            errors.append("Votes: Ballot Item %s has more votes (%d) than allowed for a question (%d)" % itemid, (c, 1))

    return errors, warnings, returndata


# Upconvert the imported data version to the current version.
def convertImportedData(data, sheetversion):
    # We know the current data set is valid, because we validated it as it was for the imported version.
    # Now we need to move that data forward to the current version for storage.
    current_user.logger.info("Checking for data conversion...", indent=1)

    # Iterate through versions until we're at the current.
    # The current version will need no conversion.
    while sheetversion < EXPORT_VERSION:
        current_user.logger.info("Converting import data version '%d' to '%d'..." % (sheetversion, (sheetversion + 1)), indent=2)

        sheetversion += 1

    current_user.logger.info("Data conversion completed.", indent=1)

    return data


# Import the previously validated data set.
def importValidatedData(data, validationdata, user):
    current_user.logger.info("Import event data: Importing validated data")

    errors = []

    # Initialize the config instance we are importing.
    imported_event = EventConfig(fetchconfig=False, version=app.config.get('VERSION'),
                                 clubid=current_user.event.clubid, eventid=current_user.event.eventid, user=user)

    # The bare minimum we need is event.
    event = data['events']
    ballotitems = data['ballotitems']
    candidates = data['candidates']

    # These others may or may not exist...
    voters = data.get('voters', None)
    votes = data.get('votes', None)

    # We need to either use or create these.
    vote_ballotid = data.get('vote_ballotid', None)

    # Reconstruct the vote ballotid.
    if vote_ballotid is None:
        vote_ballotid = []

        # Find the highest ballot value.  If there are none, the default is 0.
        # This was generated and validated during validation.
        max_ballotid = validationdata.get('max_ballotid', 0)

        # Assign that to the ballot ID.
        vote_ballotid = [{'ballotid': str(max_ballotid)}]

    # If an error occurred, stop.
    if len(errors) > 0:
        current_user.logger.flashlog("Event Data Import failure", "Data import failed:", highlight=True, large=True)

        # Limit to 25 to prevent overloading the session.  That's a lot of errors.
        for e in errors[0:24]:
            current_user.logger.flashlog("Event Data Import Failure", e, indent=True)

        if len(errors) > 25:
            current_user.logger.flashlog("Event Data Import Failure", "Too many errors to continue", indent=True)

        return redirect(url_for('main_bp.importdata'))


    # ====================================
    # Build all the INSERTs.
    # ====================================
    outsql = []

    def log_import_item(item, itemdata):
        current_user.logger.info("-> %-20s: %s" % (item, itemdata), indent=2)

    current_user.logger.debug("Importing event information...", indent=1)

    # Note that any field which is not specified is defaulted based on the EVentConfig constructor.

    for a in event:
        # Event is organized as rows of 'property' with 'value'.
        p = a['property']

        # Escpae apostrophes in these names.
        if p in ['title', 'icon', 'homeimage']:
            setattr(imported_event, p, a['value'].replace("'", "''").strip())

        elif p in ['eventdatetime']:
            # We already grabbed the date/time in the format we want, in the validation data.
            setattr(imported_event, p, validationdata['eventdatetime'])

        elif p in ['locked']:
            setattr(imported_event, p, true_false_value(a['value']))

    # The reset process will create the 'default' record for this event/club, so we need to perform an update to this record.
    outsql.append('''UPDATE events
                     SET title='%s', icon='%s', homeimage='%s', eventdatetime='%s', locked=%s
                     WHERE clubid='%d' AND eventid='%d';
                  ''' % (imported_event.title, imported_event.icon, imported_event.homeimage, imported_event.eventdatetime, imported_event.locked,
                         imported_event.clubid, imported_event.eventid))

    log_import_item("event", "%d, %d, %s, %s, %s, %s" %
                    (imported_event.clubid, imported_event.eventid, imported_event.title, imported_event.icon, imported_event.homeimage, imported_event.eventdatetime))

    current_user.logger.debug("Importing ballot items...", indent=1)
    for b in ballotitems:
        # Escape apostrophes.
        name = b['name'].replace("'", "''").strip()
        description = b['description'].replace("'", "''").strip()

        # Add each ballot item.
        outsql.append('''INSERT INTO ballotitems (clubid, eventid, itemid, type, name, description, positions, writeins)
                         VALUES('%d', '%d', '%s', '%s', '%s', '%s', '%s', '%s')
                      ''' % (imported_event.clubid, imported_event.eventid,
                             b['itemid'], b['type'], name, description, b['positions'], b['writeins']))

        log_import_item("ballotitems", "%s" % ', '.join([b['itemid'], b['type'], name, description, b['positions'], b['writeins']]))

    current_user.logger.debug("Importing candidate...", indent=1)
    for c in candidates:
        # Escape apostrophes.
        firstname = c['firstname'].replace("'", "''").strip()
        lastname = c['lastname'].replace("'", "''").strip()
        fullname = c['fullname'].replace("'", "''").strip()

        # Add each candidate.
        outsql.append('''INSERT INTO candidates (clubid, eventid, id, itemid, firstname, lastname, fullname, writein)
                         VALUES('%d', '%d', '%s', '%s', '%s', '%s', '%s', '%s')
                      ''' % (imported_event.clubid, imported_event.eventid,
                             c['id'], c['itemid'], firstname, lastname, fullname, c['writein']))

        log_import_item("candidates", "%s" % ', '.join([c['id'], c['itemid'], firstname, lastname, fullname, c['writein']]))

    if voters is not None:
        current_user.logger.debug("Importing voters...", indent=1)
        for v in voters:
            # Escape apostrophes.
            firstname = v['firstname'].replace("'", "''").strip()
            lastname = v['lastname'].replace("'", "''").strip()
            fullname = v['fullname'].replace("'", "''").strip()

            # Add each voter.
            outsql.append('''INSERT INTO voters (clubid, eventid, firstname, lastname, fullname, voteid, voted)
                            VALUES('%d', '%d', '%s', '%s', '%s', '%s', '%s')
                        ''' % (imported_event.clubid, imported_event.eventid,
                                firstname, lastname, fullname, v['voteid'], v['voted']))

            log_import_item("voters", "%s" % ', '.join([firstname, lastname, fullname, v['voteid'], v['voted']]))

    if votes is not None:
        current_user.logger.debug("Importing votes...", indent=1)
        for v in votes:
            commentary = '' if (v['commentary'] in [None, 'None'] or len(v['commentary']) == 0) else '%s' % v['commentary']

            # Add each vote.
            outsql.append('''INSERT INTO votes (clubid, eventid, itemid, ballotid, answer, commentary)
                            VALUES('%d', '%d', '%s', '%s', '%s', '%s')
                        ''' % (imported_event.clubid, imported_event.eventid,
                                v['itemid'], v['ballotid'], v['answer'], v['commentary']))

            log_import_item("votes", "%s" % ', '.join([v['itemid'], v['ballotid'], v['answer'], commentary]))

    # ====================================
    # Ballot things.
    # These will exist because we either fetched them or rebuilt them.
   # ====================================
    current_user.logger.debug("Importing ballotid data...", indent=1)
    v = vote_ballotid[0]
    imported_event.ballotid = v['ballotid']
    outsql.append('''UPDATE vote_ballotid
                    SET ballotid='%s'
                    WHERE clubid='%d' AND eventid='%d';
                ''' % (imported_event.ballotid, imported_event.clubid, imported_event.eventid))

    log_import_item("vote_ballotid", "%s" % imported_event.ballotid)

    # ====================================
    # This stuff may not exist.
    # ====================================

    # ====================================
    # Now that we've built all the queries, do the work.
    # ====================================

    # Since we are replacing data, we wipe out the system first.
    current_user.logger.info("Resetting to defaults...", indent=1, propagate=True)
    err = reset_app_data(current_user.get_userid())
    if err is not None:
        current_user.logger.flashlog("Clear Event Data failure", "Failed to reset event data:", highlight=True)
        current_user.logger.flashlog("Clear Event Data failure", err)

    # Run the transaction.
    try:
        current_user.logger.info("Importing event data...", indent=1, propagate=True)
        _, _, err = db.sql(outsql, handlekey=current_user.get_userid())
        if err is not None:
            current_user.logger.flashlog("Data Import failure", "Failed to import event data:", propagate=True)
            current_user.logger.flashlog("Data Import failure", err, propagate=True)

            return redirect(url_for('main_bp.importdata'))

    except Exception as e:
        current_user.logger.flashlog("Data Import failure", "Failed to import event data:", propagate=True)
        current_user.logger.flashlog("Data Import failure", str(e).capitalize(), propagate=True)

        return redirect(url_for('main_bp.importdata'))

    # If all went well, update the event instance.
    for attr in ['version', 'title', 'icon', 'homeimage', 'locked', 'eventdatetime']:
        setattr(current_user.event, attr, getattr(imported_event, attr))

    # Success - no URL to return.
    return None
